import dataclasses

import models

import openpyxl


def load_excel_opr(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        # Skip rows where column one is not an integer
        try:
            int(row[0])
        except:
            continue
        data.append(models.ExcelOPR(*row[:22]))
    return data


def load_excel_comment_file(file_path, vendor_model: dataclasses.dataclass):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        try:
            int(row[0])
        except:
            continue
        data.append(vendor_model(*row[:23]))
    return data


def write_merged_comments(data, filename="merged_comments.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(models.ExcelComments.header())
    for row in data:
        sheet.append(dataclasses.astuple(row))
    wb.save(filename)
